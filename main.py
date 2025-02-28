import os
import re
import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, NamedStyle
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk
import subprocess
import webbrowser
import threading

def display_banner_with_dog(text, width=50, char="*"):
    """
    Muestra un banner con un título centrado y un arte ASCII de un perro.
    
    Args:
        text (str): El título a mostrar en el banner.
        width (int): El ancho del banner.
        char (str): El carácter usado para el borde del banner.
    """
    border = char * width
    padding = (width - len(text) - 4) // 2
    centered_text = f"{char} {' ' * padding}{text}{' ' * (width - len(text) - padding - 4)}{char}"
    dog_art = """
      / \\__
    (    @\\___
    /         O
   /   (_____/
  /_____/   U
    """
    print(border)
    print(centered_text)
    print(border)
    print(dog_art)

def formatear_hoja_excel(worksheet, tipo_libro):
    """
    Ajusta el ancho de las columnas, la altura de las filas e inmoviliza la primera fila.
    
    Args:
        worksheet (Worksheet): La hoja de trabajo de Excel a formatear.
    """
    # Aplicar el estilo a las columnas específicas
    text_style = NamedStyle(name="text_style")
    text_style.font = Font(name='Arial', size=11)
    text_style.number_format = '@'  # Formato de texto

    header_style = NamedStyle(name="header_style")
    header_style.font = Font(bold=True, size=12)

    # Ajustar el ancho de las columnas
    max_width = 20  # Máximo ancho en cm
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Obtener la letra de la columna
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min((max_length + 2), max_width)
        worksheet.column_dimensions[column].width = adjusted_width

    # Ajustar la altura de las filas
    for row in worksheet.iter_rows():
        max_height = 0
        for cell in row:
            if cell.value:
                cell_height = len(str(cell.value).split('\n'))
                if cell_height > max_height:
                    max_height = cell_height
        worksheet.row_dimensions[row[0].row].height = max_height * 15

    # Inmovilizar la primera fila
    worksheet.freeze_panes = worksheet['A2']

    if tipo_libro == 'resultado':
        columnas = ['A', 'C']
    elif tipo_libro == 'actuaciones':
        columnas = ['B', 'F', 'J']
    
    for col in columnas:
        for cell in worksheet[col]:
            cell.style = text_style

    # Aplicar formato de fecha a columnas específicas si es el libro de actuaciones
    if tipo_libro == 'actuaciones':
        date_columns = ['D', 'G', 'H', 'I']
        for col in date_columns:
            for cell in worksheet[col]:
                cell.number_format = 'd-mmm-yy'
    # Convertir el contenido de las celdas de las columnas N y O en hipervínculos
    hyperlink_columns = ['M', 'N']
    for col in hyperlink_columns:
        for cell in worksheet[col]:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('https'):
                cell.hyperlink = cell.value
                cell.style = 'Hyperlink'
    # Establecer la primera fila en negrita y tamaño 12
    for cell in worksheet[1]:
        cell.style = header_style


def get_user_inputs():
    """
    Abre una ventana gráfica para solicitar al usuario la ruta del archivo, el número de columna y si desea descargar los adjuntos.
    """
    def select_file():
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls"), ("CSV files", "*.csv")])
        entry_file_path.delete(0, tk.END)
        entry_file_path.insert(0, file_path)
        if file_path.endswith('.xlsx') or file_path.endswith('.xls'):
            entry_column.config(state=tk.NORMAL)
        else:
            entry_column.config(state=tk.DISABLED)

    def submit():
        ruta_archivo = entry_file_path.get()
        numero_columna = entry_column.get()
        if ruta_archivo.endswith('.xlsx') or ruta_archivo.endswith('.xls'):
            if not numero_columna:
                messagebox.showerror("Error", "Debe ingresar el número de columna para archivos Excel.")
                return
            numero_columna = int(numero_columna) - 1
        else:
            numero_columna = None
        descargar_adjuntos = var_download.get()
        global user_inputs
        user_inputs = (ruta_archivo, numero_columna, descargar_adjuntos)
        process_file(user_inputs)

    def update_progress_bar(current, total):
        progress = (current / total) * 100
        progress_bar['value'] = progress
        progress_label.config(text=f"{int(progress)}%")
        root.update_idletasks()

    def process_file(user_inputs):
        ruta_archivo, numero_columna, descargar_adjuntos = user_inputs
        try:
            df = read_file(ruta_archivo, numero_columna)
            data = get_column_data(df)
            output_text.delete(1.0, tk.END)  # Limpiar la ventana de resultados
            progress_bar["maximum"] = len(data)
            process_data(data, descargar_adjuntos)
        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error: {e}")

    def open_file(file_path):
        """
        Abre un archivo utilizando el programa predeterminado del sistema operativo.
        
        Args:
            file_path (str): La ruta del archivo a abrir.
        """
        try:
            if os.name == 'posix':
                subprocess.call(['xdg-open', file_path])
            elif os.name == 'nt':
                os.startfile(file_path)
            else:
                messagebox.showerror("Error", "No se puede abrir el archivo en este sistema operativo.")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir el archivo: {e}")

    def open_email(event):
        webbrowser.open_new("mailto:ingmigmora@gmail.com")

    root = tk.Tk()
    root.title("Configuración de Procesos Judiciales")

    tk.Label(root, text="Ruta del archivo:").grid(row=0, column=0, padx=10, pady=10)
    entry_file_path = tk.Entry(root, width=50)
    entry_file_path.grid(row=0, column=1, padx=10, pady=10)
    tk.Button(root, text="Seleccionar archivo", command=select_file).grid(row=0, column=2, padx=10, pady=10)

    tk.Label(root, text="Número de columna (empezando desde 1):").grid(row=1, column=0, padx=10, pady=10)
    entry_column = tk.Entry(root, width=10)
    entry_column.grid(row=1, column=1, padx=10, pady=10)
    entry_column.config(state=tk.DISABLED)

    var_download = tk.StringVar(value="n")
    tk.Checkbutton(root, text="Descargar adjuntos", variable=var_download, onvalue="s", offvalue="n").grid(row=2, columnspan=3, padx=10, pady=10)

    tk.Button(root, text="Aceptar", command=submit).grid(row=3, columnspan=3, padx=10, pady=10)

    output_text = scrolledtext.ScrolledText(root, width=80, height=20)
    output_text.grid(row=4, columnspan=3, padx=10, pady=10)

    progress_bar = ttk.Progressbar(root, orient="horizontal", mode="determinate", length=640)
    progress_bar.grid(row=5, columnspan=3, padx=10, pady=10)

    progress_label = tk.Label(root, text="0%")
    progress_label.grid(row=5, column=2, padx=10, pady=10)

    developer_email = tk.Label(root, text="Contacto : ingmigmora@gmail.com", fg="green", cursor="hand2")
    developer_email.grid(row=6, columnspan=3, padx=10, pady=10)
    developer_email.bind("<Button-1>", open_email)
    
# Crear un estilo de texto
    text_style = NamedStyle(name="text_style")
    text_style.number_format = '@'  # Formato de texto

    wb_actuaciones = Workbook()
    ws_actuaciones = wb_actuaciones.active
    ws_actuaciones.title = "Actuaciones"

    wb_resultado = Workbook()
    ws_resultado = wb_resultado.active
    ws_resultado.title = "Resultado del Proceso"
    ws_resultado.append(["Número de Proceso", "Estado", "Fecha y Hora de Consulta"])  


    def print_to_output(*args):
        output_text.insert(tk.END, " ".join(map(str, args)) + "\n")
        output_text.see(tk.END)

    def process_data_actuaciones(data, rango_dias=1):
        """
        Procesa las actuaciones en los datos proporcionados y las agrupa por un rango de días especificado.

        Args:
            data (dict): Un diccionario que contiene una lista de actuaciones bajo la clave "actuaciones". 
                        Cada actuación es un diccionario con una clave "fechaActuacion" que contiene la fecha en formato ISO.
            rango_dias (int, optional): El rango de días para agrupar las actuaciones. El valor predeterminado es 1.

        Returns:
            None: La función no retorna ningún valor. Modifica directamente la hoja de cálculo `ws_actuaciones` 
                agregando las actuaciones agrupadas por el rango de días especificado.
        """
        actuaciones = data["actuaciones"]
        if not actuaciones:
            return

        # Ordenar las actuaciones por fecha
        actuaciones.sort(key=lambda x: x["fechaActuacion"], reverse=True)
        
        # Obtener la última actuación
        ultima_actuacion = actuaciones[0]
        fecha_fin = ultima_actuacion["fechaActuacion"]
        
        # Calcular la fecha de inicio
        fecha_inicio = (datetime.strptime(fecha_fin, "%Y-%m-%dT%H:%M:%S") - timedelta(days=rango_dias)).strftime("%Y-%m-%dT%H:%M:%S")
        
        # Consultar el servicio de actuaciones con las fechas especificadas
        id_proceso = data["idProceso"]
        url_actuaciones = f"https://consultaprocesos.ramajudicial.gov.co:448/api/v2/Proceso/Actuaciones/{id_proceso}?fechaIni={fecha_inicio}&fechaFin={fecha_fin}&pagina=1"
        headers = {
            "accept": "application/json, text/plain, */*",
            "user-agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
        }
        response = requests.get(url_actuaciones, headers=headers)
        
        if response.status_code != 200:
            print("Error al consultar el servicio de actuaciones")
            return
        
        actuaciones_data = response.json()
        actuaciones_filtradas = actuaciones_data.get("actuaciones", [])
        
        # Concatenar las actuaciones en la misma celda
        concatenated_actuaciones = "\n".join("\n".join(f"{key}: {value}" for key, value in actuacion.items()) for actuacion in actuaciones_filtradas)
        
        resultado = [ultima_actuacion[key] if key != "actuaciones" else concatenated_actuaciones for key in ultima_actuacion.keys()]
        ws_actuaciones.append(resultado)

    def process_data(data, descargar_adjuntos):
        """
        Procesa cada número de radicación, consulta la información del proceso y guarda los resultados en archivos Excel separados.
        
        Args:
            data (list): Lista de números de radicación.
            descargar_adjuntos (str): Indica si se deben descargar los adjuntos.
        """
        total_registros = len(data)
        exitosos = 0
        con_error = 0
        download_semaphore = threading.Semaphore(5)
        headers_written = False

        for index, numeroRadicacion in enumerate(data):
            numeroRadicacion = re.sub(r'\D', '', str(numeroRadicacion).strip()[:23])
            #numeroRadicacion = str(numeroRadicacion).strip()[:23]
            try:
                url_proceso = f"https://consultaprocesos.ramajudicial.gov.co:448/api/v2/Procesos/Consulta/NumeroRadicacion"
                params = {"numero": numeroRadicacion, "SoloActivos": "false", "pagina": "1"}
                headers = {
                    "accept": "application/json, text/plain, */*",
                    "user-agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
                }
                response = requests.get(url_proceso, headers=headers, params=params)
                proceso_data = response.json()
                update_progress_bar(index + 1, total_registros)

                if not proceso_data or not proceso_data.get('procesos'):
                    raise ValueError("No se encontró información del proceso.")
                
                procesos = proceso_data['procesos']
                for proceso in procesos:
                    id_proceso = proceso['idProceso']
                    #id_proceso = proceso_data['procesos'][0]['idProceso']
                    url_actuaciones = f"https://consultaprocesos.ramajudicial.gov.co:448/api/v2/Proceso/Actuaciones/{id_proceso}?pagina=1"
                    actuaciones_response = requests.get(url_actuaciones, headers=headers)
                    
                    if actuaciones_response.status_code == 404:
                        mensaje_error = actuaciones_response.json()["Message"]
                        raise ValueError(mensaje_error)

                    if actuaciones_response.status_code != 200 or not actuaciones_response.content:
                        raise ValueError("Respuesta inválida del servidor. consultando actuaciones. code:", actuaciones_response.status_code)


                    actuaciones_data = actuaciones_response.json()
                    if not actuaciones_data:
                        raise ValueError("No se encontraron actuaciones.")

                    ac = actuaciones_data["actuaciones"][0]

                    if not headers_written:
                        headers = list(ac.keys()) + ["URL Descarga DOC", "URL Descarga CSV", "URLs Documentos"]
                        ws_actuaciones.append(headers)
                        headers_written = True    

                    #resultado = process_data_actuaciones(actuaciones_data, rango_dias=14)
                    resultado = list(ac.values())

                    url_descarga_doc = f"https://consultaprocesos.ramajudicial.gov.co:448/api/v2/Descarga/DOCX/Proceso/{id_proceso}"
                    url_descarga_csv = f"https://consultaprocesos.ramajudicial.gov.co:448/api/v2/Descarga/CSV/Detalle/{id_proceso}"
                    resultado.append(url_descarga_doc)
                    resultado.append(url_descarga_csv)

                    if ac.get("conDocumentos") == True: 
                        id_reg_actuacion = ac["idRegActuacion"]
                        url_documentos = f"https://consultaprocesos.ramajudicial.gov.co:448/api/v2/Proceso/DocumentosActuacion/{id_reg_actuacion}"
                        headers_documentos = {
                            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8',
                            'accept-language': 'es-US,es;q=0.8',
                            'cache-control': 'max-age=0',
                            'priority': 'u=0, i',
                            'sec-ch-ua': '"Brave";v="131", "Chromium";v="131", "Not_A Brand";v="24"',
                            'sec-ch-ua-mobile': '?0',
                            'sec-ch-ua-platform': '"Linux"',
                            'sec-fetch-dest': 'document',
                            'sec-fetch-mode': 'navigate',
                            'sec-fetch-site': 'none',
                            'sec-fetch-user': '?1',
                            'sec-gpc': '1',
                            'upgrade-insecure-requests': '1',
                            'user-agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36'
                        }
                        documentos_response = requests.get(url_documentos, headers=headers_documentos)            
                        
                        if documentos_response.status_code == 200 and documentos_response.content:
                            documentos_data = documentos_response.json()
                            urls_documentos = []
                            for doc in documentos_data:
                                id_reg_documento = doc["idRegDocumento"]
                                url_descarga_documento = f"https://consultaprocesos.ramajudicial.gov.co:448/api/v2/Descarga/Documento/{id_reg_documento}"
                                urls_documentos.append(url_descarga_documento)
                            resultado.append(";".join(urls_documentos))
                            threads = []
                            if descargar_adjuntos.lower() == 's':
                                carpeta_descargas = f"./{numeroRadicacion}"
                                os.makedirs(carpeta_descargas, exist_ok=True)                                            
                                for url in urls_documentos:
                                    with download_semaphore:
                                        t = threading.Thread(
                                            target=download_file_threaded,
                                            args=(url, headers_documentos, carpeta_descargas))
                                    t.start()
                                    threads.append(t)
                            # Esperar a que todos los hilos terminen
                            for t in threads:
                                t.join()
                        else:
                            resultado.append("")
                    else:
                        resultado.append("")

                    ws_actuaciones.append(resultado)

                    fecha_hora_consulta = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    ws_resultado.append([numeroRadicacion, "Consultado correctamente", fecha_hora_consulta])
                    #ws_resultado.append([numeroRadicacion, "Consultado correctamente", fecha_hora_consulta])
                    exitosos += 1

            except ValueError as e:
                print_to_output(f"Error procesando el número de radicación {numeroRadicacion}: {e}")
                con_error += 1
                ws_resultado.append([numeroRadicacion, "Error", str(e)])
                continue  

            except Exception as e:
                ws_resultado.append([numeroRadicacion, "Error", str(e)])
                print_to_output(f"Error en el proceso {numeroRadicacion}: {e}")
                con_error += 1

            progress_bar["value"] = index + 1
            root.update_idletasks()

        try:
            formatear_hoja_excel(ws_actuaciones, "actuaciones")
            wb_actuaciones.save("actuaciones_procesos.xlsx")
            print_to_output("Archivo de actuaciones guardado exitosamente.")
        except Exception as e:
            print_to_output(f"Error al guardar el archivo de actuaciones: {e}")

        try:
            fecha_hora_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
            nombre_archivo_resultado = f"resultado_procesos_{fecha_hora_actual}.xlsx"
            formatear_hoja_excel(ws_resultado, "resultado")
            wb_resultado.save(nombre_archivo_resultado)            
            print_to_output(f"Archivo de resultados guardado exitosamente como {nombre_archivo_resultado}.")
        except Exception as e:
            print_to_output(f"Error al guardar el archivo de resultados: {e}")

        print_to_output(f"Total de registros procesados: {total_registros}")
        print_to_output(f"Registros exitosos: {exitosos}")
        print_to_output(f"Registros con error: {con_error}")

        # Botones para abrir los archivos generados
        tk.Button(root, text="Abrir archivo de actuaciones", command=lambda: open_file("actuaciones_procesos.xlsx")).grid(row=7, column=0, padx=10, pady=10)
        tk.Button(root, text="Abrir archivo de resultados", command=lambda: open_file(nombre_archivo_resultado)).grid(row=7, column=1, padx=10, pady=10)

    root.mainloop()

def read_file(ruta_archivo, numero_columna=None):
    """
    Lee un archivo Excel o CSV y devuelve un DataFrame de pandas.
    Si el archivo es de Excel, también extrae la columna especificada.
    
    Args:
        ruta_archivo (str): La ruta del archivo a leer.
        numero_columna (int, optional): El índice de la columna a extraer si el archivo es de Excel.
    
    Returns:
        DataFrame: Contiene los datos del archivo leído.
    
    Raises:
        ValueError: Si el formato del archivo no es soportado.
    """
    if ruta_archivo.endswith('.xlsx') or ruta_archivo.endswith('.xls'):
        df = pd.read_excel(ruta_archivo)
        if numero_columna is not None:
            df = df.iloc[:, [numero_columna]]
    elif ruta_archivo.endswith('.csv'):
        df = pd.read_csv(ruta_archivo)
    else:
        raise ValueError("Formato de archivo no soportado. Por favor, ingrese un archivo Excel o CSV.")
    return df

def get_column_data(df):
    """
    Obtiene los datos de la primera columna de un DataFrame y los devuelve como una lista.
    
    Args:
        df (DataFrame): El DataFrame de pandas que contiene los datos.
    
    Returns:
        list: Contiene los datos de la columna especificada.
    """
    return df.iloc[:, 0].dropna().tolist()

def download_file_threaded(url, headers, carpeta_descargas):
    """
    Descarga un archivo desde una URL y lo guarda en una carpeta específica.

    Args:
        url (str): La URL del archivo a descargar.
        headers (dict): Encabezados HTTP para la solicitud.
        carpeta_descargas (str): Carpeta donde se guardará el archivo.
    """
    try:
        archivo_response = requests.get(url, headers=headers)
        if archivo_response.status_code == 200:
            content_disposition = archivo_response.headers.get('Content-Disposition')
            if content_disposition:
                match = re.search(r'filename\*?=([^;]+)', content_disposition)
                if match:
                    nombre_archivo = match.group(1).strip().strip('"').split("''")[-1]
                else:
                    nombre_archivo = url.split("/")[-1]
            else:
                nombre_archivo = url.split("/")[-1]

            ruta_archivo = os.path.join(carpeta_descargas, nombre_archivo)
            with open(ruta_archivo, 'wb') as archivo:
                archivo.write(archivo_response.content)
    except Exception as e:
        print(f"Error descargando el archivo {url}: {e}")



def main():
    display_banner_with_dog("Consulta información de Procesos Judiciales - by Miguel M")
    get_user_inputs()

if __name__ == "__main__":
    main()
